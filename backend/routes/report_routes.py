from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from database import get_db
import models
import schemas
import auth
from datetime import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill

router = APIRouter(prefix="/api/reports", tags=["Reports"])

@router.get("/class-report")
async def generate_class_report(
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user_from_query)
):
    """Generate a PDF report for the class advisor's class with Marks and Activities"""
    
    # Verify Class Advisor role and details
    if current_user.role != models.RoleEnum.CLASS_ADVISOR or not (current_user.dept and current_user.year and current_user.section):
        raise HTTPException(status_code=403, detail="Only Class Advisors with assigned class can generate reports")

    # Determine Student Model based on Department
    dept_model_map = {
        "CSE": models.StudentCSE,
        "ECE": models.StudentECE,
        "EEE": models.StudentEEE,
        "MECH": models.StudentMECH,
        "CIVIL": models.StudentCIVIL,
        "BIO": models.StudentBIO,
        "AIDS": models.StudentAIDS
    }
    
    student_model = dept_model_map.get(current_user.dept)
    if not student_model:
        raise HTTPException(status_code=400, detail=f"Unknown department: {current_user.dept}")

    # Fetch Class Data
    students = db.query(student_model).filter(
        student_model.dept == current_user.dept,
        student_model.year == int(current_user.year),
        student_model.section == current_user.section
    ).all()
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Class Report - {current_user.dept} {current_user.year}-{current_user.section}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary Info
    summary_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>" \
                   f"Class Advisor: {current_user.name}<br/>" \
                   f"Total Students: {len(students)}"
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Student Table Data
    # Headers
    headers = ['Reg No', 'Name', 'Marks Summary', 'Activities']
    data = [headers]
    
    # Styles for table cells
    cell_style = styles['Normal']
    cell_style.fontSize = 9
    
    for student in students:
        # 1. Fetch Marks
        marks = db.query(models.Mark).filter(models.Mark.reg_no == student.reg_no).all()
        marks_summary = []
        if marks:
            for mark in marks:
                # Format marks separately
                details = []
                # Slip Tests
                st_str = f"ST:{mark.slip_test_1:g},{mark.slip_test_2:g},{mark.slip_test_3:g},{mark.slip_test_4:g}"
                details.append(st_str)
                
                # CIA
                cia_str = f"CIA:{mark.cia_1:g},{mark.cia_2:g}"
                details.append(cia_str)
                
                # Assignments
                assign_str = f"A:{mark.assignment_1:g},{mark.assignment_2:g},{mark.assignment_3:g},{mark.assignment_4:g},{mark.assignment_5:g}"
                details.append(assign_str)
                
                # Model
                details.append(f"M:{mark.model:g}")
                
                # University
                details.append(f"U:{mark.university_result_grade or '-'}")
                
                full_details = " | ".join(details)
                marks_summary.append(f"<b>{mark.subject_title}</b><br/>{full_details}")
            marks_text = "<br/><br/>".join(marks_summary)
        else:
            marks_text = "No marks found"

        # 2. Fetch Activities
        participations = db.query(models.ActivityParticipation).filter(
            models.ActivityParticipation.reg_no == student.reg_no
        ).all()
        
        activity_summary = []
        if participations:
            for p in participations:
                activity_name = p.activity.activity_name if p.activity else "Unknown Activity"
                activity_summary.append(f"• {activity_name} ({p.role or 'Participant'})")
            activity_text = "<br/>".join(activity_summary)
        else:
            activity_text = "No activities"

        # Add row to data
        data.append([
            Paragraph(student.reg_no, cell_style),
            Paragraph(student.name, cell_style),
            Paragraph(marks_text, cell_style),
            Paragraph(activity_text, cell_style)
        ])
        
    # Create Table
    col_widths = [70, 100, 200, 180]
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return Response(content=pdf_bytes, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename=class_report_{current_user.dept}_{current_user.section}.pdf"
    })

@router.get("/attendance-export")
async def export_attendance(
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user_from_query)
):
    """Export attendance data to Excel"""
    
    if current_user.role != models.RoleEnum.CLASS_ADVISOR or not (current_user.dept and current_user.year and current_user.section):
        raise HTTPException(status_code=403, detail="Only Class Advisors can export attendance")

    # Fetch Attendance Data
    attendance_records = db.query(models.Attendance).filter(
        models.Attendance.dept == current_user.dept,
        models.Attendance.year == int(current_user.year),
        models.Attendance.section == current_user.section
    ).order_by(models.Attendance.date.desc(), models.Attendance.reg_no).all()
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Log"
    
    # Headers
    headers = ["Date", "Reg No", "Name", "Status"]
    ws.append(headers)
    
    # Style Headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
    # Data
    for record in attendance_records:
        ws.append([
            record.date.strftime('%Y-%m-%d'),
            record.reg_no,
            record.student_name,
            record.status
        ])
        
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename=attendance_{current_user.dept}_{current_user.section}.xlsx"
    })

@router.get("/marks-export-excel")
async def export_class_marks_excel(
    semester: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user_from_query)
):
    """Export class marks to Excel with separate sheets per subject"""
    try:
        if current_user.role != models.RoleEnum.CLASS_ADVISOR or not (current_user.dept and current_user.year and current_user.section):
            raise HTTPException(status_code=403, detail="Only Class Advisors can export marks")

        # Determine Student Model based on Department
        dept_model_map = {
            "CSE": models.StudentCSE,
            "ECE": models.StudentECE,
            "EEE": models.StudentEEE,
            "MECH": models.StudentMECH,
            "CIVIL": models.StudentCIVIL,
            "BIO": models.StudentBIO,
            "AIDS": models.StudentAIDS
        }
        
        student_model = dept_model_map.get(current_user.dept)
        if not student_model:
            raise HTTPException(status_code=400, detail=f"Unknown department: {current_user.dept}")

        # Fetch Students
        students = db.query(student_model).filter(
            student_model.dept == current_user.dept,
            student_model.year == int(current_user.year),
            student_model.section == current_user.section
        ).order_by(student_model.reg_no).all()
        
        # Fetch all marks for the class
        reg_nos = [s.reg_no for s in students]
        query = db.query(models.Mark).filter(models.Mark.reg_no.in_(reg_nos))
        if semester:
            query = query.filter(models.Mark.semester == semester)
        
        all_marks = query.all()
        
        # Group marks by subject
        from collections import defaultdict
        marks_by_subject = defaultdict(list)
        
        for mark in all_marks:
            subject_key = f"{mark.subject_code} - {mark.subject_title}"
            marks_by_subject[subject_key].append(mark)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Style definitions
        header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        red_font = Font(color="FF0000")
        
        # Create a sheet for each subject
        for subject_key, marks_list in sorted(marks_by_subject.items()):
            # Create valid sheet name (max 31 chars, no special chars)
            sheet_name = subject_key[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
            ws = wb.create_sheet(title=sheet_name)
            
            # Headers
            headers = [
                "Reg No", "Name", 
                "ST1", "ST2", "ST3", "ST4", 
                "CIA1", "CIA2", "Model", "Uni Grade"
            ]
            ws.append(headers)
            
            # Style Headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Create a map of reg_no to mark for this subject
            marks_map = {mark.reg_no: mark for mark in marks_list}
            
            # Add data for each student
            for student in students:
                mark = marks_map.get(student.reg_no)
                
                if mark:
                    row_data = [
                        student.reg_no,
                        student.name,
                        mark.slip_test_1,
                        mark.slip_test_2,
                        mark.slip_test_3,
                        mark.slip_test_4,
                        mark.cia_1,
                        mark.cia_2,
                        mark.model,
                        mark.university_result_grade
                    ]
                    ws.append(row_data)
                    
                    # Apply Conditional Formatting
                    current_row = ws.max_row
                    
                    # Helper to check and style
                    def check_and_style(col_idx, value, threshold):
                        if value is not None and isinstance(value, (int, float)) and value < threshold:
                            ws.cell(row=current_row, column=col_idx).font = red_font
                    
                    # Slip Tests (Cols 3, 4, 5, 6) < 10
                    check_and_style(3, mark.slip_test_1, 10)
                    check_and_style(4, mark.slip_test_2, 10)
                    check_and_style(5, mark.slip_test_3, 10)
                    check_and_style(6, mark.slip_test_4, 10)
                    
                    # CIA (Cols 7, 8) < 30
                    check_and_style(7, mark.cia_1, 30)
                    check_and_style(8, mark.cia_2, 30)
                    
                    # Model (Col 9) < 50
                    check_and_style(9, mark.model, 50)
                    
                    # University Grade (Col 10) == 'ARREAR'
                    uni_cell = ws.cell(row=current_row, column=10)
                    if mark.university_result_grade == 'ARREAR':
                        uni_cell.font = red_font
        
        # If no marks found, create a single sheet with message
        if not marks_by_subject:
            ws = wb.create_sheet(title="No Marks")
            ws.append(["No marks data found for this class"])
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"marks_{current_user.dept}_{current_user.section}"
        if semester:
            filename += f"_sem{semester}"
        filename += ".xlsx"
        
        return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
            "Content-Disposition": f"attachment; filename={filename}"
        })


    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

